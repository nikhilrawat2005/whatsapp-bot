import io
import csv
import logging
import datetime
from flask import (Blueprint, request, render_template, redirect,
                   url_for, jsonify, Response, make_response)
from app.database.db import db
from app.models.models import Patient, HospitalSettings, Doctor
from app.config import Config
from app.services import slot_service
from app.services import analytics_service

logger = logging.getLogger(__name__)
web_bp = Blueprint('web', __name__)


# ──────────────────────────────────────────────────────────────────────────────
#  Public Routes
# ──────────────────────────────────────────────────────────────────────────────

@web_bp.route('/')
def home():
    """Renders the hospital landing page."""
    return render_template('index.html',
                           wa_number_id=Config.WHATSAPP_PHONE_NUMBER_ID,
                           hospital=Config,
                           departments=slot_service.DEPARTMENTS)


# ──────────────────────────────────────────────────────────────────────────────
#  Admin Dashboard
# ──────────────────────────────────────────────────────────────────────────────

@web_bp.route('/admin')
def admin_dashboard():
    """Main admin dashboard with stats, schedule, and appointment table."""
    # ── Filters (existing logic preserved) ───────────────────────────────
    search    = request.args.get('search', '').strip()
    doctor    = request.args.get('doctor', '').strip()
    date      = request.args.get('date', '').strip()
    status    = request.args.get('status', '').strip()
    date_view = request.args.get('date_view', 'today')  # today|tomorrow|week|custom

    # ── Date range calculation ────────────────────────────────────────────
    today_str    = datetime.date.today().strftime("%Y-%m-%d")
    tomorrow_str = (datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    week_start   = (datetime.date.today() - datetime.timedelta(
                    days=datetime.date.today().weekday())).strftime("%Y-%m-%d")
    week_end     = (datetime.date.today() + datetime.timedelta(
                    days=6 - datetime.date.today().weekday())).strftime("%Y-%m-%d")

    # ── Build filter query (existing logic) ───────────────────────────────
    query = Patient.query

    if search:
        query = query.filter(
            (Patient.name.like(f"%{search}%")) |
            (Patient.phone_number.like(f"%{search}%")) |
            (Patient.id == search if search.isdigit() else False)
        )
    if doctor:
        query = query.filter_by(doctor=doctor)
    if status:
        query = query.filter_by(booking_status=status)

    # Apply date_view filter unless a custom date is provided
    if date:
        query = query.filter_by(booking_date=date)
    elif date_view == 'today':
        query = query.filter_by(booking_date=today_str)
    elif date_view == 'tomorrow':
        query = query.filter_by(booking_date=tomorrow_str)
    elif date_view == 'week':
        query = query.filter(
            Patient.booking_date >= week_start,
            Patient.booking_date <= week_end
        )
    # 'all' or unrecognised: no date filter

    appointments = query.order_by(
        Patient.booking_date.asc(), Patient.booking_time.asc()
    ).all()

    # ── Stats ─────────────────────────────────────────────────────────────
    stats = analytics_service.get_dashboard_stats()

    return render_template(
        'admin.html',
        appointments=appointments,
        doctors=slot_service.DOCTORS_LIST,
        departments=slot_service.DEPARTMENTS_LIST,
        stats=stats,
        filters={
            'search': search, 'doctor': doctor,
            'date': date, 'status': status, 'date_view': date_view
        },
        today=today_str,
    )


# ──────────────────────────────────────────────────────────────────────────────
#  Appointment Actions (existing, preserved)
# ──────────────────────────────────────────────────────────────────────────────

@web_bp.route('/admin/complete/<int:appt_id>', methods=['POST'])
def complete_appointment(appt_id):
    """Marks an appointment as Completed."""
    try:
        patient = Patient.query.get(appt_id)
        if patient:
            patient.booking_status = 'Completed'
            db.session.commit()
            logger.info(f"Marked appointment ID {appt_id} as completed.")
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error completing appointment {appt_id}: {e}")
    return redirect(request.referrer or url_for('web.admin_dashboard'))


@web_bp.route('/admin/cancel/<int:appt_id>', methods=['POST'])
def cancel_appointment(appt_id):
    """Cancels appointment and frees up the slot."""
    success = slot_service.release_appointment_slot(appt_id)
    if not success:
        logger.error(f"Failed to cancel appointment {appt_id}")
    return redirect(request.referrer or url_for('web.admin_dashboard'))


@web_bp.route('/admin/confirm/<int:appt_id>', methods=['POST'])
def confirm_appointment(appt_id):
    """Confirms a Scheduled appointment."""
    try:
        patient = Patient.query.get(appt_id)
        if patient and patient.booking_status == 'Scheduled':
            patient.booking_status = 'Confirmed'
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error confirming appointment {appt_id}: {e}")
    return redirect(request.referrer or url_for('web.admin_dashboard'))


# ──────────────────────────────────────────────────────────────────────────────
#  Appointment Detail
# ──────────────────────────────────────────────────────────────────────────────

@web_bp.route('/admin/appointment/<int:appt_id>')
def appointment_detail(appt_id):
    """Detailed view of a single appointment."""
    appt = Patient.query.get_or_404(appt_id)

    # Fetch chat history for this patient
    from app.models.models import ChatMessage
    chat_history = ChatMessage.query.filter_by(patient_id=appt_id)\
        .order_by(ChatMessage.timestamp.asc()).all()

    return render_template('appointment_detail.html',
                           appt=appt, chat_history=chat_history)


@web_bp.route('/admin/appointment/<int:appt_id>/notes', methods=['POST'])
def update_notes(appt_id):
    """Update admin notes on an appointment."""
    try:
        patient = Patient.query.get_or_404(appt_id)
        patient.notes = request.form.get('notes', '').strip()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Notes update error for {appt_id}: {e}")
    return redirect(url_for('web.appointment_detail', appt_id=appt_id))


# ──────────────────────────────────────────────────────────────────────────────
#  Patient Management
# ──────────────────────────────────────────────────────────────────────────────

@web_bp.route('/admin/patients')
def patients_list():
    """Patient list with search."""
    search = request.args.get('search', '').strip()
    query  = Patient.query

    if search:
        query = query.filter(
            (Patient.name.like(f"%{search}%")) |
            (Patient.phone_number.like(f"%{search}%")) |
            (Patient.email.like(f"%{search}%") if search else False)
        )

    # Unique patients by phone number (latest record per phone)
    all_appts = query.order_by(Patient.created_at.desc()).all()

    # Group by phone number
    seen_phones = {}
    patients_grouped = []
    for appt in all_appts:
        if appt.phone_number not in seen_phones:
            seen_phones[appt.phone_number] = {
                'phone':       appt.phone_number,
                'name':        appt.name,
                'gender':      appt.gender,
                'age':         appt.age,
                'email':       appt.email,
                'last_visit':  appt.booking_date,
                'total':       1,
                'id':          appt.id,  # latest appointment id
            }
        else:
            seen_phones[appt.phone_number]['total'] += 1

    patients_grouped = list(seen_phones.values())
    return render_template('patients.html',
                           patients=patients_grouped,
                           search=search,
                           total=len(patients_grouped))


@web_bp.route('/admin/patient/<phone>')
def patient_profile(phone):
    """Individual patient profile with appointment history."""
    appointments = Patient.query.filter_by(phone_number=phone)\
        .order_by(Patient.booking_date.desc()).all()

    if not appointments:
        return redirect(url_for('web.patients_list'))

    patient_info = appointments[0]  # Most recent record for basic info

    upcoming   = [a for a in appointments if
                  a.booking_date >= datetime.date.today().strftime("%Y-%m-%d")
                  and a.booking_status in ['Scheduled', 'Confirmed']]
    past       = [a for a in appointments if
                  a.booking_status == 'Completed' or
                  a.booking_date < datetime.date.today().strftime("%Y-%m-%d")]
    cancelled  = [a for a in appointments if a.booking_status == 'Cancelled']

    return render_template('patient_profile.html',
                           patient=patient_info,
                           appointments=appointments,
                           upcoming=upcoming,
                           past=past,
                           cancelled=cancelled)


# ──────────────────────────────────────────────────────────────────────────────
#  Analytics
# ──────────────────────────────────────────────────────────────────────────────

@web_bp.route('/admin/analytics')
def analytics():
    """Analytics dashboard with charts."""
    stats        = analytics_service.get_dashboard_stats()
    weekly       = analytics_service.get_weekly_stats()
    peak         = analytics_service.get_peak_hours()
    dept_dist    = analytics_service.get_department_distribution()
    gender_dist  = analytics_service.get_gender_distribution()
    status_dist  = analytics_service.get_status_distribution()
    busiest_doc  = analytics_service.get_busiest_doctor()
    doctor_stats = analytics_service.get_doctor_stats()

    return render_template('analytics.html',
                           stats=stats,
                           weekly=weekly,
                           peak=peak,
                           dept_dist=dept_dist,
                           gender_dist=gender_dist,
                           status_dist=status_dist,
                           busiest_doc=busiest_doc,
                           doctor_stats=doctor_stats)


# ──────────────────────────────────────────────────────────────────────────────
#  Settings
# ──────────────────────────────────────────────────────────────────────────────

@web_bp.route('/admin/settings', methods=['GET'])
def settings():
    """Hospital settings management page."""
    # Load settings from DB, fallback to Config defaults
    settings_data = {
        'hospital_name':    HospitalSettings.get('hospital_name',    Config.HOSPITAL_NAME),
        'hospital_address': HospitalSettings.get('hospital_address', Config.HOSPITAL_ADDRESS),
        'hospital_phone':   HospitalSettings.get('hospital_phone',   Config.HOSPITAL_PHONE),
        'hospital_email':   HospitalSettings.get('hospital_email',   Config.HOSPITAL_EMAIL),
        'hospital_hours':   HospitalSettings.get('hospital_hours',   Config.HOSPITAL_HOURS),
        'slot_duration':    HospitalSettings.get('slot_duration',    '30'),
        'max_per_slot':     HospitalSettings.get('max_per_slot',     '1'),
        'lunch_start':      HospitalSettings.get('lunch_start',      '13:00'),
        'lunch_end':        HospitalSettings.get('lunch_end',        '14:00'),
    }
    doctors = slot_service.DOCTORS_LIST
    departments = slot_service.DEPARTMENTS_LIST
    return render_template('settings.html',
                           settings=settings_data,
                           doctors=doctors,
                           departments=departments)


@web_bp.route('/admin/settings', methods=['POST'])
def save_settings():
    """Save hospital settings."""
    fields = [
        'hospital_name', 'hospital_address', 'hospital_phone',
        'hospital_email', 'hospital_hours', 'slot_duration',
        'max_per_slot', 'lunch_start', 'lunch_end'
    ]
    try:
        for field in fields:
            value = request.form.get(field, '').strip()
            if value:
                HospitalSettings.set(field, value)
        db.session.commit()
        logger.info("Hospital settings saved successfully.")
    except Exception as e:
        db.session.rollback()
        logger.error(f"Settings save error: {e}")

    return redirect(url_for('web.settings'))


# ──────────────────────────────────────────────────────────────────────────────
#  Export
# ──────────────────────────────────────────────────────────────────────────────

@web_bp.route('/admin/export/csv')
def export_csv():
    """Export filtered appointments as CSV."""
    search = request.args.get('search', '').strip()
    doctor = request.args.get('doctor', '').strip()
    date   = request.args.get('date', '').strip()
    status = request.args.get('status', '').strip()

    query = Patient.query
    if search:
        query = query.filter(
            (Patient.name.like(f"%{search}%")) |
            (Patient.phone_number.like(f"%{search}%"))
        )
    if doctor:
        query = query.filter_by(doctor=doctor)
    if date:
        query = query.filter_by(booking_date=date)
    if status:
        query = query.filter_by(booking_status=status)

    appointments = query.order_by(Patient.booking_date.asc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', 'Name', 'Age', 'Gender', 'Phone', 'Email',
        'Doctor', 'Department', 'Date', 'Time', 'Status', 'Problem', 'Notes', 'Created At'
    ])
    for a in appointments:
        writer.writerow([
            a.id, a.name, a.age, a.gender, a.phone_number, a.email or '',
            a.doctor, a.department or '', a.booking_date, a.booking_time,
            a.booking_status, a.problem or '', a.notes or '',
            a.created_at.strftime('%Y-%m-%d %H:%M') if a.created_at else ''
        ])

    output.seek(0)
    filename = f"appointments_{datetime.date.today().strftime('%Y%m%d')}.csv"
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@web_bp.route('/admin/export/excel')
def export_excel():
    """Export appointments as Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "openpyxl not installed. Run: pip install openpyxl", 500

    appointments = Patient.query.order_by(Patient.booking_date.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    headers = [
        'ID', 'Name', 'Age', 'Gender', 'Phone', 'Email',
        'Doctor', 'Department', 'Date', 'Time', 'Status', 'Problem', 'Notes'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, a in enumerate(appointments, 2):
        ws.append([
            a.id, a.name, a.age, a.gender, a.phone_number, a.email or '',
            a.doctor, a.department or '', a.booking_date, a.booking_time,
            a.booking_status, a.problem or '', a.notes or ''
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"appointments_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    response = make_response(buf.getvalue())
    response.headers['Content-Type'] = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response
